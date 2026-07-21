"""
PQ.ai Performance Dashboard — Flask app.

Reads the latest workbook uploaded to GCS by report.py/main_job.py
(gs://<bucket>/<reports_prefix>/*.xlsx) and serves an interactive dashboard:
KPI cards (from the Summary sheet) plus, per bucket table (ACV_Bucket,
Sale_Price_Bucket, Title_Type_Bucket, AutoGrade_Bucket, Loss_Type_Bucket,
Lot_Type_Bucket), two side-by-side charts:
    1. Percent-metrics chart: PQ / PQ_ai Mean Error Pct and MAPE (Cleansed)
    2. Value-metrics chart: Units Sold and ASP - PQ / PQ_ai (Cleansed)

Usage:
    python dashboard.py
    -> open http://127.0.0.1:8080

Config (env vars):
    GCS_BUCKET_NAME     — required. Bucket the reports are uploaded to.
    GCS_REPORTS_PREFIX  — default "reports/pq_ai_weekly_report/"
    BQ_KEY_PATH         — optional. Local service-account JSON for GCS auth.
                          On Cloud Run, leave unset — Application Default
                          Credentials (the runtime service account) is used.

ASSUMPTIONS (verify / adjust to match your actual "Summary (Sheet 1)" query
columns — I only had report.py's formatting logic + your screenshots to infer
these, not the real column names):
    - Summary sheet has a "breakout" column with values matching the detail
      sheet labels ("BluCar ex-TFSS", "Insurance + TFSS") and a "period"
      column with values in {"Past Week", "Past Month", "Trailing 3 Months"}.
    - Summary sheet has columns: Units Sold, PQ Coverage, PQ_ai Coverage,
      PQ Error %, PQ_ai Error %, PQ MAE, PQ_ai MAE, PQ MAPE, PQ_ai MAPE.
    - If your actual column names differ, tweak KPI_CARD_SPECS below —
      the app does a case-insensitive substring match so minor differences
      ("PQai" vs "PQ_ai") are usually fine, but wildly different names won't be.
    - Detail-sheet "Cleansed" columns (per bucket sub-table) are:
        "PQ Mean Error Pct - Cleansed", "PQ_ai Mean Error Pct - Cleansed",
        "PQ MAPE - Cleansed", "PQ_ai MAPE - Cleansed",
        "Units Sold", "ASP - PQ Cleansed", "ASP - PQ_ai Cleansed"
      All %-type columns are assumed to be Excel percent-format fractions
      (e.g. 0.029 -> 2.9%) and are scaled by x100 for display. Worth a
      spot-check against a raw cell value the first time this runs against
      real data.
"""

import os
import tempfile
from flask import Flask, render_template_string, jsonify, request
from openpyxl import load_workbook

from gcs_utils import get_latest_blob_name, download_blob_to_file

pwd = os.path.dirname(os.path.abspath(__file__))

GCS_BUCKET = os.environ.get("GCS_BUCKET_NAME")
GCS_REPORTS_PREFIX = os.environ.get("GCS_REPORTS_PREFIX", "reports/pq_ai_weekly_report/")
KEY_PATH = os.environ.get("BQ_KEY_PATH")  # None on Cloud Run -> uses ADC

# Local-folder fallback: if this directory exists and contains an .xlsx file,
# use the most recently modified one instead of hitting GCS. Handy for local
# dev/testing against output/pq_ai_weekly_report/ without needing bucket access.
# Set LOCAL_REPORTS_DIR="" to force GCS even if the folder exists.
LOCAL_REPORTS_DIR = os.environ.get(
    "LOCAL_REPORTS_DIR", os.path.join(pwd, "output", "pq_ai_weekly_report")
)

SUMMARY_SHEET = "PQ.ai Summary"
DETAIL_SHEETS = {
    "BluCar ex-TFSS": "BluCar ex-TFSS PQ.ai Detail",
    "Insurance + TFSS": "Insurance + TFSS PQ.ai Detail",
}
PERIODS = ["Past Week", "Past Month", "Trailing 3 Months"]

# Fixed column layout for every bucket sub-table in the detail sheets, confirmed
# directly against the source workbook (columns D, E, G, I, L, N are hidden
# spacer/raw columns and are intentionally skipped). Column A = period label,
# column B = bucket value. 1-indexed to match openpyxl's ws.cell(row, col).
#   C=3: Units Sold          F=6:  ASP - PQ Cleansed
#   H=8: PQ Mean Error Pct   J=10: PQ MAPE - Cleansed
#   K=11: ASP - PQ_ai        M=13: PQ_ai Mean Error Pct
#   O=15: PQ_ai MAPE
DETAIL_METRIC_COLUMNS = {
    "units": 3,                  # C — Units Sold
    "pq_asp": 6,                 # F — ASP - PQ Cleansed
    "pq_mean_error_pct": 8,      # H — PQ Mean Error Pct - Cleansed
    "pq_mape": 10,                # J — PQ MAPE - Cleansed
    "pqai_asp": 11,               # K — ASP - PQ_ai Cleansed
    "pqai_mean_error_pct": 13,   # M — PQ_ai Mean Error Pct - Cleansed
    "pqai_mape": 15,              # O — PQ_ai MAPE - Cleansed
}

KPI_CARD_SPECS = [
    ("Units Sold", "units sold", None),
    ("PQ Coverage", "pq coverage", "pq"),
    ("PQ_ai Coverage", "pq_ai coverage", "pqai"),
    ("PQ Error %", "pq error", "pq"),
    ("PQ_ai Error %", "pq_ai error", "pqai"),
    ("PQ MAE", "pq mae", "pq"),
    ("PQ_ai MAE", "pq_ai mae", "pqai"),
    ("PQ MAPE", "pq mape", "pq"),
    ("PQ_ai MAPE", "pq_ai mape", "pqai"),
]

app = Flask(__name__)

# Simple in-memory cache: avoid re-downloading from GCS on every single
# request. Keyed on the blob's last-updated timestamp — if a newer report
# lands in GCS, the cache naturally invalidates on the next check.
_cache = {"blob_name": None, "updated": None, "local_path": None}


def _find_latest_local_xlsx():
    """Return path to the most recently modified .xlsx in LOCAL_REPORTS_DIR, or None.
    Disabled on Cloud Run (K_SERVICE set) so a stray .xlsx accidentally baked into
    the image can never shadow the real GCS report in production."""
    if "K_SERVICE" in os.environ:
        return None
    if not LOCAL_REPORTS_DIR or not os.path.isdir(LOCAL_REPORTS_DIR):
        return None
    candidates = [
        os.path.join(LOCAL_REPORTS_DIR, f)
        for f in os.listdir(LOCAL_REPORTS_DIR)
        if f.lower().endswith(".xlsx") and not f.startswith("~$")
    ]
    if not candidates:
        return None
    return max(candidates, key=os.path.getmtime)


def _fetch_latest_workbook_path():
    # Prefer a local file if LOCAL_REPORTS_DIR has one — skips GCS entirely.
    local_path = _find_latest_local_xlsx()
    if local_path:
        mtime = os.path.getmtime(local_path)
        if _cache["blob_name"] == local_path and _cache["updated"] == mtime and _cache["local_path"]:
            return _cache["local_path"]
        _cache.update({"blob_name": local_path, "updated": mtime, "local_path": local_path})
        return local_path

    if not GCS_BUCKET:
        raise RuntimeError(
            "GCS_BUCKET_NAME is not set and no local report was found in "
            f"'{LOCAL_REPORTS_DIR}'. The dashboard needs either a local .xlsx "
            "there or a GCS bucket to read reports from."
        )

    blob_name, updated = get_latest_blob_name(GCS_BUCKET, GCS_REPORTS_PREFIX, key_path=KEY_PATH)

    if _cache["blob_name"] == blob_name and _cache["updated"] == updated and _cache["local_path"]:
        return _cache["local_path"]

    # New/updated report — download fresh, replacing any previous temp file.
    if _cache["local_path"] and os.path.exists(_cache["local_path"]):
        os.remove(_cache["local_path"])

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    download_blob_to_file(GCS_BUCKET, blob_name, tmp.name, key_path=KEY_PATH)

    _cache.update({"blob_name": blob_name, "updated": updated, "local_path": tmp.name})
    return tmp.name


def _find_col(columns, needle):
    """Case-insensitive substring match; returns the matching column name or None."""
    needle = needle.lower()
    for c in columns:
        if needle in str(c).lower():
            return c
    return None


def _as_pct(val):
    """Excel percent-format cells store fractions (0.029 -> 2.9%); scale to display %."""
    if val is None:
        return None
    try:
        return float(val) * 100
    except (TypeError, ValueError):
        return None


def _as_num(val):
    if val is None:
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _parse_summary(ws):
    """Row 1 = headers, rows 2+ = data (matches _update_summary in report.py)."""
    headers = []
    for cell in ws[1]:
        if cell.value is None:
            break
        headers.append(cell.value)

    rows = []
    for row in ws.iter_rows(min_row=2, max_col=len(headers)):
        if all(c.value is None for c in row):
            continue
        rows.append({headers[i]: row[i].value for i in range(len(headers))})
    return headers, rows


def _parse_detail_sheet(ws):
    """
    Reverse of _write_sub_table / _update_detail_sheet in report.py:
    each sub-table is: header row (col B = bucket name), then for each
    period: a period row (col A only), then data rows (col B = bucket
    value, metric values at the FIXED columns in DETAIL_METRIC_COLUMNS).
    A blank row separates tables.

    NOTE: metric values are read by fixed column index (see
    DETAIL_METRIC_COLUMNS), not by scanning/matching header text. The sheet
    has hidden spacer/raw-data columns (D, E, G, I, L, N) interleaved with
    the "Cleansed" columns we want — a contiguous header scan breaks on
    those, so we go straight to the known-good column letters instead.

    Returns: dict of { bucket_col_name: { period: [ {bucket, <metric_key>: val, ...}, ... ] } }
    """
    tables = {}
    max_row = ws.max_row
    row_idx = 3  # data starts at row 3 in detail sheets

    while row_idx <= max_row:
        col_a = ws.cell(row_idx, 1).value
        col_b = ws.cell(row_idx, 2).value

        if col_a is None and col_b is None:
            row_idx += 1
            continue

        # Header row for a new sub-table: col B holds the bucket label
        # (e.g. "ACV_Bucket"). No need to scan metric headers — columns
        # are fixed via DETAIL_METRIC_COLUMNS.
        bucket_col = col_b
        row_idx += 1

        table_data = {p: [] for p in PERIODS}
        current_period = None

        while row_idx <= max_row:
            col_a = ws.cell(row_idx, 1).value
            col_b = ws.cell(row_idx, 2).value

            if col_a is None and col_b is None:
                row_idx += 1
                break  # blank gap -> next sub-table

            if col_a is not None and col_b is None:
                current_period = col_a
                if current_period not in table_data:
                    table_data[current_period] = []
                row_idx += 1
                continue

            # data row — read each metric from its fixed column
            entry = {"bucket": col_b}
            for key, col_idx in DETAIL_METRIC_COLUMNS.items():
                entry[key] = ws.cell(row_idx, col_idx).value
            if current_period is not None:
                table_data[current_period].append(entry)
            row_idx += 1

        tables[bucket_col] = table_data

    return tables


def load_dashboard_data():
    path = _fetch_latest_workbook_path()
    wb = load_workbook(path, data_only=True)

    summary_headers, summary_rows = _parse_summary(wb[SUMMARY_SHEET])

    detail_data = {}
    for label, sheet_name in DETAIL_SHEETS.items():
        if sheet_name in wb.sheetnames:
            detail_data[label] = _parse_detail_sheet(wb[sheet_name])

    wb.close()
    return {
        "file": os.path.basename(_cache["blob_name"]) if _cache["blob_name"] else os.path.basename(path),
        "summary_headers": summary_headers,
        "summary_rows": summary_rows,
        "detail_data": detail_data,
    }


def get_kpis(summary_headers, summary_rows, breakout_label, period):
    breakout_col = _find_col(summary_headers, "breakout") or _find_col(summary_headers, "segment")
    period_col = _find_col(summary_headers, "period")

    match = None
    for row in summary_rows:
        b_ok = (breakout_col is None) or (
            str(row.get(breakout_col, "")).strip().lower() == breakout_label.strip().lower()
        )
        p_ok = (period_col is None) or (
            str(row.get(period_col, "")).strip().lower() == period.strip().lower()
        )
        if b_ok and p_ok:
            match = row
            break

    kpis = []
    for label, needle, group in KPI_CARD_SPECS:
        col = _find_col(summary_headers, needle)
        val = match.get(col) if (match and col) else None
        kpis.append({"label": label, "value": _fmt_value(label, val), "group": group})
    return kpis


def _fmt_value(label, val):
    if val is None:
        return "—"
    try:
        f = float(val)
    except (TypeError, ValueError):
        return str(val)

    lower = label.lower()
    if "%" in label or "coverage" in lower or "mape" in lower or "error" in lower:
        return f"{f * 100:.1f}%" if abs(f) <= 1.5 else f"{f:.1f}%"
    if "units" in lower:
        return f"{f:,.0f}"
    return f"${f:,.0f}"


@app.route("/")
def index():
    return render_template_string(PAGE_TEMPLATE, sheets=list(DETAIL_SHEETS.keys()), periods=PERIODS)


@app.route("/healthz")
def healthz():
    """Basic liveness check for Cloud Run."""
    return jsonify({"status": "ok"})


@app.route("/api/debug")
def api_debug():
    """
    TEMPORARY debug route — remove once verified against source workbook.
    Dumps, for a given sheet/period/bucket, the raw values read from the
    fixed columns in DETAIL_METRIC_COLUMNS, so you can diff against the
    source .xlsx directly.
    Usage: /api/debug?sheet=BluCar ex-TFSS&period=Past Month&bucket=ACV_Bucket
    """
    sheet_label = request.args.get("sheet", list(DETAIL_SHEETS.keys())[0])
    period = request.args.get("period", PERIODS[1])
    bucket = request.args.get("bucket")

    data = load_dashboard_data()
    detail = data["detail_data"].get(sheet_label, {})

    if bucket is None:
        return jsonify({"available_buckets": list(detail.keys())})

    by_period = detail.get(bucket, {})
    rows = by_period.get(period, [])
    if not rows:
        return jsonify({"error": f"No rows for bucket={bucket!r} period={period!r}",
                         "available_periods": list(by_period.keys())})

    return jsonify({
        "column_mapping_used": DETAIL_METRIC_COLUMNS,
        "raw_rows": rows,  # exact bucket + every metric value as read from the fixed columns
    })


@app.route("/api/data")
def api_data():
    sheet_label = request.args.get("sheet", list(DETAIL_SHEETS.keys())[0])
    period = request.args.get("period", PERIODS[1])

    try:
        data = load_dashboard_data()
    except FileNotFoundError as e:
        return jsonify({"error": str(e)}), 404

    kpis = get_kpis(data["summary_headers"], data["summary_rows"], sheet_label, period)

    charts = {}
    detail = data["detail_data"].get(sheet_label, {})
    for bucket_col, by_period in detail.items():
        rows = by_period.get(period, [])
        if not rows:
            continue

        charts[bucket_col] = {
            "labels": [r["bucket"] for r in rows],
            "pq_mean_error_pct": [_as_pct(r["pq_mean_error_pct"]) for r in rows],
            "pqai_mean_error_pct": [_as_pct(r["pqai_mean_error_pct"]) for r in rows],
            "pq_mape": [_as_pct(r["pq_mape"]) for r in rows],
            "pqai_mape": [_as_pct(r["pqai_mape"]) for r in rows],
            "units": [_as_num(r["units"]) for r in rows],
            "pq_asp": [_as_num(r["pq_asp"]) for r in rows],
            "pqai_asp": [_as_num(r["pqai_asp"]) for r in rows],
        }

    return jsonify({"file": data["file"], "kpis": kpis, "charts": charts})


PAGE_TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<title>PQ.ai Performance Dashboard</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
<style>
  * { box-sizing: border-box; }
  body { font-family: -apple-system, Segoe UI, Roboto, Arial, sans-serif; margin: 0; background: #f4f6f8; color: #1a1a2e; }
  header { background: #16213e; color: #fff; padding: 20px 32px; }
  header h1 { margin: 0; font-size: 22px; }
  header p { margin: 4px 0 0; color: #b8c1d9; font-size: 13px; }
  .controls { display: flex; gap: 16px; padding: 16px 32px; background: #fff; border-bottom: 1px solid #e2e6ea; align-items: center; }
  .controls label { font-weight: 600; font-size: 13px; margin-right: 6px; }
  select { padding: 6px 10px; border-radius: 6px; border: 1px solid #ccc; font-size: 13px; }
  .kpis { display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr)); gap: 14px; padding: 20px 32px; }
  .kpi-card { background: #fff; border-radius: 8px; padding: 14px 16px; box-shadow: 0 1px 3px rgba(0,0,0,.08); border-top: 4px solid #d8dde3; }
  .kpi-card.pq { border-top-color: #4472C4; }
  .kpi-card.pqai { border-top-color: #70AD47; }
  .kpi-label { font-size: 11px; letter-spacing: .04em; color: #6b7280; text-transform: uppercase; font-weight: 700; }
  .kpi-value { font-size: 24px; font-weight: 700; margin-top: 4px; }
  .charts { display: grid; grid-template-columns: repeat(2, 1fr); gap: 20px; padding: 0 32px 32px; }
  @media (max-width: 900px) { .charts { grid-template-columns: 1fr; } }
  .chart-card { background: #fff; border-radius: 8px; padding: 16px; box-shadow: 0 1px 3px rgba(0,0,0,.08); }
  .chart-card h3 { margin: 0 0 10px; font-size: 15px; }
  .chart-card canvas { margin-bottom: 16px; }
  .chart-card canvas:last-child { margin-bottom: 0; }
  .meta { padding: 0 32px; font-size: 12px; color: #6b7280; }
  .error { padding: 24px 32px; color: #b91c1c; }
</style>
</head>
<body>
<header>
  <h1>PQ.ai Performance Dashboard</h1>
  <p id="subtitle">Loading...</p>
</header>
<div class="controls">
  <div>
    <label for="sheetSelect">Segment:</label>
    <select id="sheetSelect">
      {% for s in sheets %}<option value="{{ s }}">{{ s }}</option>{% endfor %}
    </select>
  </div>
  <div>
    <label for="periodSelect">Period:</label>
    <select id="periodSelect">
      {% for p in periods %}<option value="{{ p }}" {% if p == "Past Month" %}selected{% endif %}>{{ p }}</option>{% endfor %}
    </select>
  </div>
</div>
<div id="errorBanner" class="error" style="display:none;"></div>
<div class="kpis" id="kpiContainer"></div>
<div class="charts" id="chartContainer"></div>
<div class="meta" id="fileMeta"></div>

<script>
const chartInstances = {};

async function loadData() {
  const sheet = document.getElementById('sheetSelect').value;
  const period = document.getElementById('periodSelect').value;
  const res = await fetch(`/api/data?sheet=${encodeURIComponent(sheet)}&period=${encodeURIComponent(period)}`);
  const data = await res.json();

  const errorBanner = document.getElementById('errorBanner');
  if (data.error) {
    errorBanner.style.display = 'block';
    errorBanner.textContent = 'No report found yet: ' + data.error;
    document.getElementById('kpiContainer').innerHTML = '';
    document.getElementById('chartContainer').innerHTML = '';
    return;
  }
  errorBanner.style.display = 'none';

  document.getElementById('subtitle').textContent = `${sheet} | ${period}`;
  document.getElementById('fileMeta').textContent = `Source: ${data.file}`;

  const kpiContainer = document.getElementById('kpiContainer');
  kpiContainer.innerHTML = '';
  data.kpis.forEach(k => {
    const div = document.createElement('div');
    div.className = 'kpi-card' + (k.group ? ' ' + k.group : '');
    div.innerHTML = `<div class="kpi-label">${k.label}</div><div class="kpi-value">${k.value}</div>`;
    kpiContainer.appendChild(div);
  });

  const chartContainer = document.getElementById('chartContainer');
  chartContainer.innerHTML = '';
  Object.entries(data.charts).forEach(([bucketName, c]) => {
    const safeId = bucketName.replace(/[^a-zA-Z0-9]/g, '_');
    const cardIdPct = 'chart_pct_' + safeId;
    const cardIdUnits = 'chart_units_' + safeId;
    const cardIdAsp = 'chart_asp_' + safeId;

    const card = document.createElement('div');
    card.className = 'chart-card';
    card.innerHTML = `<h3>${bucketName}</h3>
      <canvas id="${cardIdPct}"></canvas>
      <canvas id="${cardIdUnits}"></canvas>
      <canvas id="${cardIdAsp}"></canvas>`;
    chartContainer.appendChild(card);

    // Chart 1: percent metrics (Mean Error Pct + MAPE, PQ vs PQ_ai)
    const ctxPct = document.getElementById(cardIdPct).getContext('2d');
    if (chartInstances[cardIdPct]) chartInstances[cardIdPct].destroy();
    chartInstances[cardIdPct] = new Chart(ctxPct, {
      data: {
        labels: c.labels,
        datasets: [
          { type: 'bar', label: 'PQ Mean Error Pct', data: c.pq_mean_error_pct, backgroundColor: '#4472C4' },
          { type: 'bar', label: 'PQ_ai Mean Error Pct', data: c.pqai_mean_error_pct, backgroundColor: '#9DC3E6' },
          { type: 'bar', label: 'PQ MAPE', data: c.pq_mape, backgroundColor: '#548235' },
          { type: 'bar', label: 'PQ_ai MAPE', data: c.pqai_mape, backgroundColor: '#A9D18E' },
        ]
      },
      options: {
        responsive: true,
        interaction: { mode: 'index', intersect: false },
        scales: {
          y: { title: { display: true, text: 'Percent (%)' } },
        }
      }
    });

    // Chart 2: Units Sold (its own scale, no ASP mixed in)
    const ctxUnits = document.getElementById(cardIdUnits).getContext('2d');
    if (chartInstances[cardIdUnits]) chartInstances[cardIdUnits].destroy();
    chartInstances[cardIdUnits] = new Chart(ctxUnits, {
      type: 'bar',
      data: {
        labels: c.labels,
        datasets: [
          { label: 'Units Sold', data: c.units, backgroundColor: '#d9d9d9' },
        ]
      },
      options: {
        responsive: true,
        interaction: { mode: 'index', intersect: false },
        scales: {
          y: { title: { display: true, text: 'Units' } },
        }
      }
    });

    // Chart 3: ASP - PQ vs PQ_ai (its own scale, no Units mixed in)
    const ctxAsp = document.getElementById(cardIdAsp).getContext('2d');
    if (chartInstances[cardIdAsp]) chartInstances[cardIdAsp].destroy();
    chartInstances[cardIdAsp] = new Chart(ctxAsp, {
      type: 'line',
      data: {
        labels: c.labels,
        datasets: [
          { label: 'ASP - PQ', data: c.pq_asp, borderColor: '#1F3864', tension: 0.2 },
          { label: 'ASP - PQ_ai', data: c.pqai_asp, borderColor: '#375623', tension: 0.2 },
        ]
      },
      options: {
        responsive: true,
        interaction: { mode: 'index', intersect: false },
        scales: {
          y: { title: { display: true, text: 'ASP ($)' } },
        }
      }
    });
  });
}

document.getElementById('sheetSelect').addEventListener('change', loadData);
document.getElementById('periodSelect').addEventListener('change', loadData);
loadData();
</script>
</body>
</html>
"""

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8080))
    # Cloud Run sets PORT/K_SERVICE automatically on deploy — use that as the
    # signal to run in production mode (debug off). Locally, neither is set,
    # so debug=True (auto-reload, interactive debugger) is used by default.
    is_cloud_run = "K_SERVICE" in os.environ
    app.run(host="0.0.0.0", port=port, debug=not is_cloud_run)