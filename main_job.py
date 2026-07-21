import pandas as pd
import numpy as np
import json
import datetime
import argparse
from operator import itemgetter

import smtplib
from os.path import basename
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import COMMASPACE, formatdate

import logging
import os
import inspect
import sys
import re
import zipfile
import io
import shutil
import math
from concurrent.futures import ThreadPoolExecutor, as_completed
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from email.mime.image import MIMEImage
from google.cloud import storage as gcs_storage

# try:
#     pwd = os.path.abspath(os.path.join(os.path.dirname(__file__), "../../"))
# except:
#     pwd = ('/opt/copart/pentaho/data')

pwd = os.path.abspath(os.path.join(os.path.dirname(__file__), "../../"))
desktop_config = '/Users/srdeo/Desktop/config'
if desktop_config not in sys.path:
    sys.path.insert(0, desktop_config)
sys.path.append(f'{pwd}/config')

from include.initialize import initialize

jobs_definition_path = f'{pwd}/resources/jobs.json'
app_env = 'prod'


def set_logger(dir_path):
    logger = logging.getLogger(__name__)
    logger.setLevel(logging.INFO)

    file_handler = logging.FileHandler(os.path.join(dir_path, f'{job_name}.log'))
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
    logger.addHandler(file_handler)

    return logger


def connect_to_database(env, database):
    print(f'Connecting to {database}...')
    conn = initialize(env, database)
    try:
        connection = conn.get_connection()
        print(f'Connected to {database}...')

        return connection

    except Exception as e:
        print(f"Error connecting to {database}: {e}")
        logger.error(f"Error connecting to {database}: {e}")
        raise


def get_data(query, database):
    with open(f"{pwd}/resources/queries/{query}", "r") as f:
        sql_command = f.read()

    connection = connect_to_database(app_env, database)

    try:
        print(f'Fetching data from {database}...')
        df = connection.query(sql_command).to_dataframe()
        connection.close()
        print(f'Fetched data from {database}...')
        print(f"{df.shape[0]} rows of data received from {database}")

        if not df.empty and len(df):
            last_row = df.iloc[-1].to_dict()
            logger.info(last_row)
        else:
            logger.error('Data Frame Empty / Could not get last row')

        return df

    except Exception as e:
        print(f"Error getting data from {database}: {e}")
        logger.error(f"Error getting data from {database} for {job_name}: {e}")
        raise


def sort_files_by_last_modified(files):
    """ Given a list of files, return them sorted by the last
         modified times. """
    fileData = {}
    for fname in files:
        fileData[fname] = os.stat(fname).st_mtime

    fileData = sorted(fileData.items(), key=itemgetter(1))
    return fileData


def delete_oldest_files(sorted_files, keep):
    """ Given a list of files sorted by last modified time and a number to
        keep, delete the oldest ones. """
    delete = len(sorted_files) - keep
    for x in range(0, delete):
        os.remove(sorted_files[x][0])


def remove_old_files(output_dir_path, file_type='.csv', keep=30):
    """ Given a path to directory list all the files that end with output file type
    , sort files list, keep latest 30 files. """
    dir_files = []
    for dir, sub_dir, files in os.walk(output_dir_path, topdown=False):
        for name in files:
            if name.endswith(file_type):
                dir_files.append(os.path.join(dir, name))

    sorted_dir_files = sort_files_by_last_modified(dir_files)

    delete_oldest_files(sorted_dir_files, keep=keep)


def send_mail(send_from, send_to, subject, text, files=None,
              server="smtp.copart.com", inline_images=None):
    """
    send_from: sender's email id
    send_to: reciever's email id or list of email id
    subject: text for subject line
    text: plain text for body of email (also used as html fallback)
    files: location of files you want to attach
    inline_images: list of image file paths to embed inline in the email body
    """
    try:
        assert isinstance(send_to, list)

        msg = MIMEMultipart("related")
        msg['From'] = send_from
        msg['To'] = COMMASPACE.join(send_to)
        msg['Date'] = formatdate(localtime=True)
        msg['Subject'] = subject

        html_body = text.replace("\n", "<br>")
        if inline_images:
            for i, img_path in enumerate(inline_images):
                cid = f"chart_{i}"
                html_body += f'<br><br><img src="cid:{cid}" style="max-width:100%;">'

        alt_part = MIMEMultipart("alternative")
        alt_part.attach(MIMEText(text, "plain"))
        alt_part.attach(MIMEText(f"<html><body>{html_body}</body></html>", "html"))
        msg.attach(alt_part)

        for f in files or []:
            with open(f, "rb") as fil:
                part = MIMEApplication(fil.read(), Name=basename(f))
            part['Content-Disposition'] = 'attachment; filename="%s"' % basename(f)
            msg.attach(part)

        for i, img_path in enumerate(inline_images or []):
            with open(img_path, "rb") as img_file:
                img_part = MIMEImage(img_file.read())
            img_part.add_header("Content-ID", f"<chart_{i}>")
            img_part.add_header("Content-Disposition", "inline", filename=basename(img_path))
            msg.attach(img_part)

        smtp = smtplib.SMTP(server, 587)
        smtp.sendmail(send_from, send_to, msg.as_string())
        smtp.close()

    except Exception as e:
        print(f"Error sending email: {e}")
        logger.error(f"Error sending email: {e}")
        raise


def environ_or_required(key):
    return (
        {'default': os.environ.get(key)} if os.environ.get(key) else {'required': True}
    )


# ── Mean Error Pct threshold alerting ───────────────────────────────────────
# If PQ Mean Error Pct - Cleansed or PQ_ai Mean Error Pct - Cleansed on the
# "Past Month" Summary row (for either breakout: BluCar ex-TFSS or
# Insurance + TFSS) exceeds this threshold in either direction, an alert
# email is sent. This runs unconditionally (regardless of --test), right
# after the GCS upload, since this job's host is already trusted by the
# smtp.copart.com relay — unlike e.g. a laptop on VPN, which gets silently
# dropped/timed-out by the relay's IP allow-list.
ALERT_THRESHOLD_PCT = float(os.environ.get("ALERT_THRESHOLD_PCT", "0.03"))  # 3%
ALERT_PERIOD_FILTER = os.environ.get("ALERT_PERIOD_FILTER", "Past Month")
ALERT_SMTP_SERVER = os.environ.get("ALERT_SMTP_SERVER", "smtp.copart.com")
ALERT_FROM_EMAIL = os.environ.get("ALERT_FROM_EMAIL", "")
ALERT_TO_EMAILS = [e.strip() for e in os.environ.get("ALERT_TO_EMAILS", "").split(",") if e.strip()]


def _find_summary_breaches(df_summary, period_filter=ALERT_PERIOD_FILTER):
    """
    Check df_summary (the "Summary (Sheet 1)" dataframe, same one written to
    the Summary sheet) for rows matching `period_filter` (both breakouts)
    where abs(PQ Mean Error Pct - Cleansed) or
    abs(PQ_ai Mean Error Pct - Cleansed) exceeds ALERT_THRESHOLD_PCT.
    Returns a list of dicts: {breakout, period, metric, value}.
    """
    cols = list(df_summary.columns)

    def _find_col(needle, exclude_ai=False):
        needle_l = needle.lower()
        for c in cols:
            cl = str(c).lower()
            if exclude_ai and ("pq_ai" in cl or "pqai" in cl):
                continue
            if needle_l in cl:
                return c
        return None

    breakout_col = _find_col("breakout") or _find_col("segment")
    period_col = _find_col("period")
    pq_col = _find_col("mean error pct - cleansed", exclude_ai=True) or _find_col("mean error pct", exclude_ai=True)
    pqai_col = _find_col("pq_ai mean error pct")

    breaches = []
    for _, row in df_summary.iterrows():
        period_val = row.get(period_col, "—") if period_col else "—"
        if period_filter and str(period_val).strip().lower() != period_filter.strip().lower():
            continue

        breakout_val = row.get(breakout_col, "—") if breakout_col else "—"

        for label, col in (("PQ Mean Error Pct - Cleansed", pq_col),
                            ("PQ_ai Mean Error Pct - Cleansed", pqai_col)):
            if not col:
                continue
            val = row.get(col)
            try:
                val_f = float(val)
            except (TypeError, ValueError):
                continue
            if abs(val_f) > ALERT_THRESHOLD_PCT:
                breaches.append({
                    "breakout": breakout_val,
                    "period": period_val,
                    "metric": label,
                    "value": val_f,
                })
    return breaches


def _send_alert_email(breaches, report_label):
    """Send a plain-text alert email listing every threshold breach found."""
    if not ALERT_FROM_EMAIL or not ALERT_TO_EMAILS:
        print("Alert email skipped: ALERT_FROM_EMAIL / ALERT_TO_EMAILS not configured.")
        return

    lines = [
        f"PQ.ai Mean Error Pct exceeded ±{ALERT_THRESHOLD_PCT * 100:.1f}% "
        f"in the latest report ({report_label}):",
        "",
    ]
    for b in breaches:
        lines.append(
            f"  - {b['breakout']} | {b['period']} | {b['metric']}: {b['value'] * 100:.1f}%"
        )
    body = "\n".join(lines)

    msg = MIMEMultipart()
    msg["From"] = ALERT_FROM_EMAIL
    msg["To"] = COMMASPACE.join(ALERT_TO_EMAILS)
    msg["Date"] = formatdate(localtime=True)
    msg["Subject"] = f"[PQ.ai Alert] Mean Error Pct threshold breached — {report_label}"
    msg.attach(MIMEText(body, "plain"))

    try:
        smtp = smtplib.SMTP(ALERT_SMTP_SERVER, 587)
        smtp.sendmail(ALERT_FROM_EMAIL, ALERT_TO_EMAILS, msg.as_string())
        smtp.close()
        print(f"Alert email sent for {report_label} ({len(breaches)} breach(es)).")
    except Exception as e:
        print(f"Alert email failed to send: {e}")
        logger.error(f"Alert email failed to send: {e}")


def check_and_alert_summary(df_summary, report_label):
    """Run the threshold check against df_summary and email if breached."""
    breaches = _find_summary_breaches(df_summary)
    if breaches:
        _send_alert_email(breaches, report_label)
    else:
        print(f"No Mean Error Pct threshold breaches for {report_label} "
              f"(period={ALERT_PERIOD_FILTER!r}, threshold={ALERT_THRESHOLD_PCT * 100:.1f}%).")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--job_name", default='')
    parser.add_argument("--current_run_date", default='')
    parser.add_argument("--test", default="False")
    parser.add_argument("--user", default="paulzhi")
    args = parser.parse_args()
    job_name = args.job_name
    current_run_date = args.current_run_date
    test = args.test
    test_user = args.user

    if not os.path.exists(f'{pwd}/output/{job_name}'):
        os.makedirs(f'{pwd}/output/{job_name}')

    log_path = os.path.dirname(os.path.realpath(f"{pwd}/output/logs/{job_name}.log"))
    logger = set_logger(log_path)
    with open(jobs_definition_path, 'r') as f:
        jobs_json = json.load(f)
    jobs = jobs_json['jobs']
    job = jobs[job_name]

    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.views import Selection

    # ── CONFIG ─────────────────────────────────────────────────────────────────

    TEMPLATE_PATH = f'{pwd}/resources/scripts/PQ_ai_Reporting_Reference_Outputs.xlsx'
    SQL_PATH      = f'{pwd}/resources/queries/{job["query"]}'

    SUMMARY_SHEET  = "PQ.ai Summary"
    BLUCAR_SHEET   = "BluCar ex-TFSS PQ.ai Detail"
    INS_TFSS_SHEET = "Insurance + TFSS PQ.ai Detail"

    _SHEET_RENAMES = {
        "PQ VQ Summary":              SUMMARY_SHEET,
        "BluCar ex-TFSS VQ Detail":   BLUCAR_SHEET,
        "Insurance + TFSS VQ Detail": INS_TFSS_SHEET,
    }

    _SECTION_NAMES = (
        "Summary (Sheet 1)",
        "PQ.ai Error ACV Bucket View - BluCar ex-TFSS",
        "PQ.ai Error Sale Price Bucket View - BluCar ex-TFSS",
        "PQ.ai Error Loss Type - BluCar ex-TFSS",
        "PQ.ai Error Lot Type - BluCar ex-TFSS",
        "PQ.ai Error AutoGrade Bucket - BluCar ex-TFSS",
        "PQ.ai Error Title Type - BluCar ex-TFSS",
        "PQ.ai Error ACV Bucket View - Insurance + TFSS",
        "PQ.ai Error Sale Price Bucket View - Insurance + TFSS",
        "PQ.ai Error Loss Type - Insurance + TFSS",
        "PQ.ai Error Lot Type - Insurance + TFSS",
        "PQ.ai Error Make Bucket - Insurance + TFSS",
        "PQ.ai Error Title Type - Insurance + TFSS",
    )

    _DOLLAR_FMT = '_("$"* #,##0_);_("$"* \\(#,##0\\);_("$"* "-"??_);_(@_)'
    _PCT_FMT    = "0.0%"
    _PERIODS    = ["Past Week", "Past Month", "Trailing 3 Months"]

    GCS_DASHBOARD_BUCKET = os.environ.get("GCS_DASHBOARD_BUCKET", "cprtqa-sads-pqai-dashboard")

    _SECTION_TO_FILENAME = {
        "Summary (Sheet 1)":                                       "summary",
        "PQ.ai Error ACV Bucket View - Insurance + TFSS":          "ins_tfss_acv_bucket",
        "PQ.ai Error Sale Price Bucket View - Insurance + TFSS":   "ins_tfss_sale_price_bucket",
        "PQ.ai Error Title Type - Insurance + TFSS":               "ins_tfss_title_type",
        "PQ.ai Error Make Bucket - Insurance + TFSS":              "ins_tfss_make_bucket",
        "PQ.ai Error Loss Type - Insurance + TFSS":                "ins_tfss_loss_type",
        "PQ.ai Error Lot Type - Insurance + TFSS":                 "ins_tfss_lot_type",
        "PQ.ai Error ACV Bucket View - BluCar ex-TFSS":            "blucar_acv_bucket",
        "PQ.ai Error Sale Price Bucket View - BluCar ex-TFSS":     "blucar_sale_price_bucket",
        "PQ.ai Error Title Type - BluCar ex-TFSS":                 "blucar_title_type",
        "PQ.ai Error AutoGrade Bucket - BluCar ex-TFSS":           "blucar_autograde_bucket",
        "PQ.ai Error Loss Type - BluCar ex-TFSS":                  "blucar_loss_type",
        "PQ.ai Error Lot Type - BluCar ex-TFSS":                   "blucar_lot_type",
    }

    _FILL_PQ        = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")  # light blue
    _FILL_PQAI      = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # light green
    _FILL_PQAI_LOW  = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # light orange
    _FILL_PQAI_HIGH = PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid")  # light purple

    def _col_fill(col_name):
        if "PQ_ai Low" in col_name:
            return _FILL_PQAI_LOW
        if "PQ_ai High" in col_name:
            return _FILL_PQAI_HIGH
        if "PQ_ai" in col_name or "ProQuote_ai" in col_name:
            return _FILL_PQAI
        if "PQ " in col_name or "ProQuote" in col_name:
            return _FILL_PQ
        return None

    def _upload_to_gcs(dfs_dict, run_date, bucket_name=GCS_DASHBOARD_BUCKET):
        """Upload all section DataFrames as Parquet to GCS for the Cloud Run dashboard to read."""
        # Same prod/pii_access service-account key the usmart BigQuery connection
        # uses (include/initialize.py) — this environment has no ADC available.
        json_file_path = '/opt/airflow/dags/data/config/include/cprtpr-datastewards-sp1-614d7e297848.json'
        client = gcs_storage.Client.from_service_account_json(json_file_path)
        bucket = client.bucket(bucket_name)

        metadata = {
            "run_date": run_date,
            "run_ts": datetime.datetime.now(datetime.timezone.utc).isoformat(),
            "section_count": len(dfs_dict),
        }

        def _upload_blob(prefix, name, data, content_type):
            blob = bucket.blob(f"{prefix}/{name}")
            blob.upload_from_string(data, content_type=content_type)

        upload_jobs = []
        for section_name, df in dfs_dict.items():
            safe_name = _SECTION_TO_FILENAME.get(section_name)
            if safe_name is None:
                print(f"  Skipping unknown section for GCS upload: {section_name}")
                continue
            parquet_bytes = df.to_parquet(index=False)
            for prefix in ("latest", f"archive/{run_date}"):
                upload_jobs.append((prefix, f"{safe_name}.parquet", parquet_bytes, "application/octet-stream", safe_name))

        metadata_bytes = json.dumps(metadata, indent=2).encode("utf-8")
        for prefix in ("latest", f"archive/{run_date}"):
            upload_jobs.append((prefix, "_metadata.json", metadata_bytes, "application/json", "_metadata"))

        with ThreadPoolExecutor(max_workers=min(len(upload_jobs), 8)) as executor:
            futures = {
                executor.submit(_upload_blob, prefix, name, data, content_type): label
                for prefix, name, data, content_type, label in upload_jobs
            }
            for future in as_completed(futures):
                future.result()
                print(f"  Uploaded {futures[future]}")

    # ── PARSE SQL ──────────────────────────────────────────────────────────────

    with open(SQL_PATH, encoding='utf-8') as f:
        sql_content = f.read()

    _pattern = "|".join(re.escape(n) for n in _SECTION_NAMES)
    _parts   = re.split(rf"(?m)^--\s*({_pattern})\s*$", sql_content)

    queries = {}
    for i in range(1, len(_parts), 2):
        name  = _parts[i].strip()
        query = _parts[i + 1].strip() if i + 1 < len(_parts) else ""
        queries[name] = query

    print(f"SQL sections loaded: {list(queries.keys())}")

    # ── RUN ALL 9 BIGQUERY QUERIES (in parallel — each section gets its own connection) ─
    def _run_section_query(section_name, sql):
        print(f"Running '{section_name}'...")
        conn = connect_to_database(app_env, job['database'])
        df = conn.query(sql).to_dataframe()
        conn.close()
        print(f"  → {section_name}: {len(df):,} rows, {len(df.columns)} cols")
        return section_name, df

    dfs = {}
    with ThreadPoolExecutor(max_workers=min(len(queries), 8)) as executor:
        futures = [executor.submit(_run_section_query, name, sql) for name, sql in queries.items()]
        for future in as_completed(futures):
            section_name, df = future.result()
            dfs[section_name] = df

    # Restore original section order (as_completed returns results out of order)
    dfs = {name: dfs[name] for name in queries if name in dfs}

    _DROP_COLS = ["PQ Cleansed Units Sold", "PQ_ai Cleansed Units Sold"]
    for key in dfs:
        dfs[key] = dfs[key].drop(columns=[c for c in _DROP_COLS if c in dfs[key].columns])

    df2 = dfs["Summary (Sheet 1)"]

    # ── EXCEL HELPER FUNCTIONS ─────────────────────────────────────────────────

    def _xl(v):
        """Convert pandas NA/NaT to None so openpyxl can write the cell."""
        try:
            if pd.isna(v):
                return None
        except (TypeError, ValueError):
            pass
        return v

    def _col_fmt(col_name):
        """Return the Excel number format for a metric column."""
        if "Variance" in col_name or "Pct" in col_name or col_name.startswith("%") or "MAPE" in col_name:
            return _PCT_FMT
        if col_name in ("Volume", "Units Sold"):
            return "#,##0"
        if col_name in ("breakout", "period"):
            return "General"
        return _DOLLAR_FMT

    def _update_summary(ws, df):
        """Overwrite Summary sheet: write headers to row 1, clear rows 2+, paste all data."""
        ws.delete_rows(1, ws.max_row)

        df_cols = list(df.columns)
        bold = Font(bold=True)
        col_fills = [_col_fill(c) for c in df_cols]

        for ci, col_name in enumerate(df_cols, start=1):
            cell = ws.cell(1, ci)
            cell.value = col_name
            cell.font  = bold
            if col_fills[ci - 1]:
                cell.fill = col_fills[ci - 1]

        for i, (_, row_data) in enumerate(df.iterrows()):
            excel_row = 2 + i
            for ci, col_name in enumerate(df_cols, start=1):
                cell = ws.cell(excel_row, ci)
                cell.value         = _xl(row_data[col_name])
                cell.number_format = _col_fmt(col_name)

        print(f"  Summary done – {len(df):,} rows written.")

    def _write_sub_table(ws, df, start_row):
        """Write one sub-table (ACV Bucket / Loss Type / etc.) with period subheaders.

        Returns the first row after all written content.
        """
        df_cols     = list(df.columns)   # [period, bucket_col, metric1, …]
        period_col  = df_cols[0]
        bucket_col  = df_cols[1]
        metric_cols = df_cols[2:]
        met_fills   = [_col_fill(mc) for mc in metric_cols]

        bold = Font(bold=True)
        row  = start_row

        # Table header row
        ws.cell(row, 2).value = bucket_col
        ws.cell(row, 2).font  = bold
        for ci, mc in enumerate(metric_cols, start=3):
            cell = ws.cell(row, ci)
            cell.value = mc
            cell.font  = bold
            if met_fills[ci - 3]:
                cell.fill = met_fills[ci - 3]
        row += 1

        # One group per period
        for period in _PERIODS:
            period_df = df[df[period_col] == period]
            if period_df.empty:
                continue

            ws.cell(row, 1).value = period
            ws.cell(row, 1).font  = bold
            row += 1

            for _, row_data in period_df.iterrows():
                ws.cell(row, 2).value = _xl(row_data[bucket_col])
                for ci, mc in enumerate(metric_cols, start=3):
                    cell               = ws.cell(row, ci)
                    cell.value         = _xl(row_data[mc])
                    cell.number_format = _col_fmt(mc)
                row += 1

        return row

    def _update_detail_sheet(ws, tables):
        """Clear rows 3+ and write all sub-tables with period subheaders."""
        if ws.max_row >= 3:
            ws.delete_rows(3, ws.max_row - 2)

        current_row = 3
        for idx, df in enumerate(tables):
            current_row = _write_sub_table(ws, df, current_row)
            if idx < len(tables) - 1:
                current_row += 1  # blank gap row between sub-tables

        print(f"  Detail sheet done – last data row: {current_row - 1}.")

    # ── BUILD OUTPUT WORKBOOK ──────────────────────────────────────────────────

    output_filename = f"{pwd}/output/{job_name}/{job_name}_{current_run_date}.{job['output_file_type']}"
    shutil.copy2(TEMPLATE_PATH, output_filename)
    print(f"Template copied → {output_filename}")

    wb = load_workbook(output_filename)

    for old_name, new_name in _SHEET_RENAMES.items():
        if old_name in wb.sheetnames:
            wb[old_name].title = new_name

    print(f"Updating '{SUMMARY_SHEET}'...")
    _update_summary(wb[SUMMARY_SHEET], df2)

    print(f"Updating '{BLUCAR_SHEET}'...")
    _update_detail_sheet(wb[BLUCAR_SHEET], [
        dfs["PQ.ai Error ACV Bucket View - BluCar ex-TFSS"],
        dfs["PQ.ai Error Sale Price Bucket View - BluCar ex-TFSS"],
        dfs["PQ.ai Error Title Type - BluCar ex-TFSS"],
        dfs["PQ.ai Error AutoGrade Bucket - BluCar ex-TFSS"],
        dfs["PQ.ai Error Loss Type - BluCar ex-TFSS"],
        dfs["PQ.ai Error Lot Type - BluCar ex-TFSS"],
    ])

    print(f"Updating '{INS_TFSS_SHEET}'...")
    _update_detail_sheet(wb[INS_TFSS_SHEET], [
        dfs["PQ.ai Error ACV Bucket View - Insurance + TFSS"],
        dfs["PQ.ai Error Sale Price Bucket View - Insurance + TFSS"],
        dfs["PQ.ai Error Title Type - Insurance + TFSS"],
        dfs["PQ.ai Error Make Bucket - Insurance + TFSS"],
        dfs["PQ.ai Error Loss Type - Insurance + TFSS"],
        dfs["PQ.ai Error Lot Type - Insurance + TFSS"],
    ])

    for ws in wb.worksheets:
        # Reset to a single fresh Selection — strips stale bottomLeft/bottomRight
        # entries inherited from the template (which would trigger Excel's "View"
        # repair against the new xSplit-only pane), while leaving the one entry
        # that openpyxl's freeze_panes setter expects to populate.
        ws.sheet_view.selection = [Selection()]
        ws.freeze_panes = "C1"
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            col_letter = get_column_letter(col_idx)
            for row_idx in range(1, min(ws.max_row + 1, 50)):
                val = ws.cell(row_idx, col_idx).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

    wb.save(output_filename)
    wb.close()
    print(f"Workbook saved.")

    # ── UPLOAD TO GCS FOR CLOUD RUN DASHBOARD ──────────────────────────────────
    try:
        _upload_to_gcs(dfs, current_run_date or datetime.datetime.now().strftime("%Y-%m-%d"))
        print("GCS upload complete.")
    except Exception as e:
        print(f"GCS upload failed (non-fatal): {e}")
        logger.error(f"GCS upload failed: {e}")

    # ── MEAN ERROR PCT THRESHOLD ALERT ──────────────────────────────────────────
    # Runs unconditionally (not gated by --test), right after the GCS upload,
    # since this host is already trusted by the smtp.copart.com relay.
    try:
        check_and_alert_summary(df2, report_label=os.path.basename(output_filename))
    except Exception as e:
        print(f"Threshold alert check failed (non-fatal): {e}")
        logger.error(f"Threshold alert check failed: {e}")

    # openpyxl rewrites xl/drawings/drawing1.xml with a default namespace
    # instead of the template's xdr:* prefix, which triggers Excel's
    # "Repaired Records: Drawing shape" on open. Restore the part
    # byte-for-byte from the template, preserving per-entry compression.
    _DRAWING_PARTS = ("xl/drawings/drawing1.xml",)
    with zipfile.ZipFile(TEMPLATE_PATH, "r") as tz:
        template_blobs = {p: tz.read(p) for p in _DRAWING_PARTS if p in tz.namelist()}
    if template_blobs:
        patched_path = output_filename + ".patched"
        with zipfile.ZipFile(output_filename, "r") as zin, \
             zipfile.ZipFile(patched_path, "w") as zout:
            for item in zin.infolist():
                data = template_blobs.get(item.filename, zin.read(item.filename))
                zout.writestr(item, data)
        shutil.move(patched_path, output_filename)
        print("Drawing part restored from template.")

    # ── GENERATE CHART ────────────────────────────────────────────────────────
    df_ins_sale = dfs["PQ.ai Error Sale Price Bucket View - Insurance + TFSS"]
    period_col = df_ins_sale.columns[0]
    bucket_col = df_ins_sale.columns[1]
    metric_cols = [c for c in ["PQ Mean Error Pct - Cleansed", "PQ MAE - Cleansed"] if c in df_ins_sale.columns]

    periods = ["Past Week", "Past Month", "Trailing 3 Months"]
    buckets = df_ins_sale[df_ins_sale[period_col] == "Past Week"][bucket_col].tolist()

    n_metrics = len(metric_cols)
    fig, axes = plt.subplots(n_metrics, 1, figsize=(14, 5 * n_metrics))
    if n_metrics == 1:
        axes = [axes]

    colors = ["#4C72B0", "#DD8452", "#55A868"]
    bar_width = 0.25
    x = np.arange(len(buckets))

    for ax, metric in zip(axes, metric_cols):
        for i, period in enumerate(periods):
            period_df = df_ins_sale[df_ins_sale[period_col] == period].reset_index(drop=True)
            values = period_df[metric].tolist() if not period_df.empty else [0] * len(buckets)
            ax.bar(x + i * bar_width, values, width=bar_width, label=period, color=colors[i])
        ax.set_title(f"Insurance + TFSS — {metric}", fontsize=12, fontweight="bold")
        ax.set_xticks(x + bar_width)
        ax.set_xticklabels(buckets, rotation=30, ha="right", fontsize=9)
        ax.legend(fontsize=9)
        ax.grid(axis="y", linestyle="--", alpha=0.5)
        is_pct = "Pct" in metric or "MAPE" in metric or "Variance" in metric
        if is_pct:
            ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f"{v:.1%}"))

    plt.suptitle("Insurance + TFSS PQ.ai — Sale Price Bucket", fontsize=14, fontweight="bold", y=1.01)
    plt.tight_layout()

    chart_filename = f"{pwd}/output/{job_name}/{job_name}_{current_run_date}_chart.png"
    plt.savefig(chart_filename, dpi=150, bbox_inches="tight")
    plt.close()
    print(f"Chart saved → {chart_filename}")

    # ── GENERATE COMBO (VOLUME BAR + ERROR LINE) CHARTS FOR REMAINING BUCKETS ──
    def _generate_bucket_combo_chart(df, bucket_label, chart_path):
        """Bars = Units Sold (Trailing 3 Months, primary axis).
        Lines = error metric per period (secondary axis).
        Shows whether a high-error bucket is backed by real volume or is just noise."""
        period_col = df.columns[0]
        bucket_col = df.columns[1]
        metric_cols = [c for c in ["PQ Mean Error Pct - Cleansed", "PQ MAE - Cleansed"] if c in df.columns]
        volume_col = "Units Sold" if "Units Sold" in df.columns else None

        volume_period = "Trailing 3 Months"
        buckets_local = df[df[period_col] == volume_period][bucket_col].tolist()
        if not buckets_local:
            for p in periods:
                candidate = df[df[period_col] == p][bucket_col].tolist()
                if candidate:
                    buckets_local = candidate
                    volume_period = p
                    break

        x_local = np.arange(len(buckets_local))
        n_metrics_local = len(metric_cols)

        fig, axes = plt.subplots(n_metrics_local, 1, figsize=(14, 5.5 * n_metrics_local))
        if n_metrics_local == 1:
            axes = [axes]

        vol_df = df[df[period_col] == volume_period].set_index(bucket_col)
        volumes = [vol_df[volume_col].get(b, 0) if volume_col else 0 for b in buckets_local]

        for ax, metric in zip(axes, metric_cols):
            ax.bar(x_local, volumes, width=0.5, color="#D9D9D9", label=f"Units Sold ({volume_period})", zorder=1)
            ax.set_ylabel("Units Sold", fontsize=10)
            ax.set_xticks(x_local)
            ax.set_xticklabels(buckets_local, rotation=30, ha="right", fontsize=9)
            ax.grid(axis="y", linestyle="--", alpha=0.3, zorder=0)

            ax2 = ax.twinx()
            for i, period in enumerate(periods):
                period_df = df[df[period_col] == period].set_index(bucket_col)
                values = [period_df[metric].get(b, np.nan) if not period_df.empty else np.nan for b in buckets_local]
                ax2.plot(x_local, values, marker="o", linewidth=2, label=period, color=colors[i], zorder=3)

            is_pct = "Pct" in metric or "MAPE" in metric or "Variance" in metric
            if is_pct:
                ax2.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f"{v:.1%}"))
            ax2.set_ylabel(metric, fontsize=10)
            ax.set_title(f"{bucket_label} — {metric}", fontsize=12, fontweight="bold")

            bars_h, bars_l = ax.get_legend_handles_labels()
            lines_h, lines_l = ax2.get_legend_handles_labels()
            ax.legend(bars_h + lines_h, bars_l + lines_l, fontsize=8, loc="upper right")

        plt.suptitle(f"Insurance + TFSS PQ.ai — {bucket_label}", fontsize=14, fontweight="bold", y=1.01)
        plt.tight_layout()
        plt.savefig(chart_path, dpi=150, bbox_inches="tight")
        plt.close()
        print(f"Chart saved → {chart_path}")

    _BUCKET_CHART_SOURCES = {
        "ACV Bucket":       "PQ.ai Error ACV Bucket View - Insurance + TFSS",
        "Title Type Bucket": "PQ.ai Error Title Type - Insurance + TFSS",
        "Make Bucket":      "PQ.ai Error Make Bucket - Insurance + TFSS",
        "Loss Type Bucket": "PQ.ai Error Loss Type - Insurance + TFSS",
        "Lot Type Bucket":  "PQ.ai Error Lot Type - Insurance + TFSS",
    }

    bucket_chart_filenames = []
    for bucket_label, df_key in _BUCKET_CHART_SOURCES.items():
        if df_key not in dfs:
            continue
        safe_label = bucket_label.replace(" ", "_")
        bucket_chart_path = f"{pwd}/output/{job_name}/{job_name}_{current_run_date}_chart_{safe_label}.png"
        _generate_bucket_combo_chart(dfs[df_key], bucket_label, bucket_chart_path)
        bucket_chart_filenames.append(bucket_chart_path)

    all_chart_filenames = [chart_filename] + bucket_chart_filenames

    if test == "True":
        name_email_map = {#'owen': 'owen.swetenburg@copart.com',
                          #'aashish': 'aashish.khadka@copart.com',
                          #'john': 'john.paul@copart.com',
                          #'pushkaraj': 'pushkaraj.jadhav@copart.com',
                          #'paulzhi': 'ziyun.zhi@copart.com',
                          'rohit': 'rohith.kokkula@copart.com',
                          'srdeo': 'srijan.deo@copart.com'}  # Ensure that you add your name in the python script to test it
        SERVER = "smtp.copart.com"
        # FROM = job['from']
        try:
            FROM = name_email_map[test_user]
        except KeyError:
            test_user = 'pushkaraj'
            FROM = name_email_map[test_user]
        # TO = job['to']
        TO = [name_email_map[test_user], job["author"]]
        SUBJECT = job['subject']
        if (not df2.empty and len(
                df2)):  # 'or (not df3.empty and len(df3))' -> you can add this code before ':' to modify if there more than 2 tabs in reference excel
            output_filename = f"{pwd}/output/{job_name}/{job_name}_{current_run_date}.{job['output_file_type']}"
            TEXT = f"Hi {'/'.join([address.split('.')[0].capitalize() for address in TO])},\n\n{job['text']}\n\nReport Owner: {job['author']}"
            FILES = [output_filename]

        else:
            TEXT = f"Hi {'/'.join([address.split('.')[0].capitalize() for address in TO])},\n\nThere are no records for this iteration of report.\n\nReport Owner: {job['author']}"
            FILES = None
        try:
            send_mail(FROM, TO, SUBJECT, TEXT, FILES, SERVER, inline_images=all_chart_filenames)
        except Exception as e:
            print(f"Email send failed: {e}")
        remove_old_files(f"{pwd}/output/{job_name}", file_type=job['output_file_type'], keep=job['keep'])
        print(f"job {job_name} completed on {datetime.datetime.now()}")

    else:
        output_filename = f"{pwd}/output/{job_name}/{job_name}.{job['output_file_type']}"
        remove_old_files(f"{pwd}/output/{job_name}", file_type=job['output_file_type'], keep=job['keep'])

    v_py_exec_success = 'True'