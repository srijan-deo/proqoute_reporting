"""
PQ.ai reporting job — BigQuery to Excel only.

Usage from main.py:

    from report import run_report

    output_path = run_report(
        job_name="pq_ai_weekly_report",
        current_run_date="2026-07-20",
        key_path="/path/to/service_account.json",
    )

Standalone (CLI) usage is still supported — see the __main__ block at the
bottom.
"""

import pandas as pd
import datetime
import argparse
from operator import itemgetter

import os
import re
import zipfile
import shutil
from google.cloud import bigquery

pwd = os.path.dirname(os.path.abspath(__file__))


def get_bq_client(key_path):
    """Build a BigQuery client directly from a service-account JSON key file."""
    print('Connecting to BigQuery...')
    try:
        client = bigquery.Client.from_service_account_json(key_path)
        print('Connected to BigQuery...')
        return client
    except Exception as e:
        print(f"Error connecting to BigQuery: {e}")
        raise


def sort_files_by_last_modified(files):
    """ Given a list of files, return them sorted by the last
         modified times. """
    fileData = {}
    for fname in files:
        fileData[fname] = os.stat(fname).st_mtime

    fileData = sorted(fileData.items(), key=itemgetter(1))
    return fileData


def delete_oldest_files(sorted_files, keep):
    """ Given a list of files sorted by last modified time and a number to
        keep, delete the oldest ones. """
    delete = len(sorted_files) - keep
    for x in range(0, delete):
        os.remove(sorted_files[x][0])


def remove_old_files(output_dir_path, file_type='.csv', keep=30):
    """ Given a path to directory list all the files that end with output file type
    , sort files list, keep latest 30 files. """
    dir_files = []
    for dir, sub_dir, files in os.walk(output_dir_path, topdown=False):
        for name in files:
            if name.endswith(file_type):
                dir_files.append(os.path.join(dir, name))

    sorted_dir_files = sort_files_by_last_modified(dir_files)

    delete_oldest_files(sorted_dir_files, keep=keep)


def run_report(
    job_name,
    current_run_date,
    key_path,
    query_filename="report.sql",
    output_file_type="xlsx",
    keep=30,
):
    """
    Runs the PQ.ai report job: pulls BigQuery data and builds the Excel
    workbook. Returns the path to the saved workbook.

    key_path: path to the GCP service-account JSON key, used for BigQuery auth.
    query_filename: name of the .sql file (in the project root) containing
        the labelled query sections.
    output_file_type: extension for the output workbook (e.g. "xlsx").
    keep: number of old output files to retain when cleaning up.
    """

    if not os.path.exists(f'{pwd}/output/{job_name}'):
        os.makedirs(f'{pwd}/output/{job_name}')

    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.views import Selection

    # ── CONFIG ─────────────────────────────────────────────────────────────────

    TEMPLATE_PATH = f'{pwd}/PQ_ai_Reporting_Reference_Outputs.xlsx'
    SQL_PATH      = f'{pwd}/{query_filename}'

    SUMMARY_SHEET  = "PQ.ai Summary"
    BLUCAR_SHEET   = "BluCar ex-TFSS PQ.ai Detail"
    INS_TFSS_SHEET = "Insurance + TFSS PQ.ai Detail"

    _SHEET_RENAMES = {
        "PQ VQ Summary":              SUMMARY_SHEET,
        "BluCar ex-TFSS VQ Detail":   BLUCAR_SHEET,
        "Insurance + TFSS VQ Detail": INS_TFSS_SHEET,
    }

    _SECTION_NAMES = (
        "Summary (Sheet 1)",
        "PQ.ai Error ACV Bucket View - BluCar ex-TFSS",
        "PQ.ai Error Sale Price Bucket View - BluCar ex-TFSS",
        "PQ.ai Error Loss Type - BluCar ex-TFSS",
        "PQ.ai Error Lot Type - BluCar ex-TFSS",
        "PQ.ai Error AutoGrade Bucket - BluCar ex-TFSS",
        "PQ.ai Error Title Type - BluCar ex-TFSS",
        "PQ.ai Error ACV Bucket View - Insurance + TFSS",
        "PQ.ai Error Sale Price Bucket View - Insurance + TFSS",
        "PQ.ai Error Loss Type - Insurance + TFSS",
        "PQ.ai Error Lot Type - Insurance + TFSS",
        "PQ.ai Error Make Bucket - Insurance + TFSS",
        "PQ.ai Error Title Type - Insurance + TFSS",
    )

    _DOLLAR_FMT = '_("$"* #,##0_);_("$"* \\(#,##0\\);_("$"* "-"??_);_(@_)'
    _PCT_FMT    = "0.0%"
    _PERIODS    = ["Past Week", "Past Month", "Trailing 3 Months"]

    _FILL_PQ        = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")  # light blue
    _FILL_PQAI      = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # light green
    _FILL_PQAI_LOW  = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # light orange
    _FILL_PQAI_HIGH = PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid")  # light purple

    def _col_fill(col_name):
        if "PQ_ai Low" in col_name:
            return _FILL_PQAI_LOW
        if "PQ_ai High" in col_name:
            return _FILL_PQAI_HIGH
        if "PQ_ai" in col_name or "ProQuote_ai" in col_name:
            return _FILL_PQAI
        if "PQ " in col_name or "ProQuote" in col_name:
            return _FILL_PQ
        return None

    # ── PARSE SQL ──────────────────────────────────────────────────────────────

    with open(SQL_PATH, encoding='utf-8') as f:
        sql_content = f.read()

    _pattern = "|".join(re.escape(n) for n in _SECTION_NAMES)
    _parts   = re.split(rf"(?m)^--\s*({_pattern})\s*$", sql_content)

    queries = {}
    for i in range(1, len(_parts), 2):
        name  = _parts[i].strip()
        query = _parts[i + 1].strip() if i + 1 < len(_parts) else ""
        queries[name] = query

    print(f"SQL sections loaded: {list(queries.keys())}")

    # ── RUN ALL 9 BIGQUERY QUERIES ─────────────────────────────────────────────

    bq_client = get_bq_client(key_path)

    dfs = {}
    for section_name, sql in queries.items():
        print(f"Running '{section_name}'...")
        dfs[section_name] = bq_client.query(sql).to_dataframe()
        print(f"  → {len(dfs[section_name]):,} rows, {len(dfs[section_name].columns)} cols")

    _DROP_COLS = ["PQ Cleansed Units Sold", "PQ_ai Cleansed Units Sold"]
    for key in dfs:
        dfs[key] = dfs[key].drop(columns=[c for c in _DROP_COLS if c in dfs[key].columns])

    df2 = dfs["Summary (Sheet 1)"]

    # ── EXCEL HELPER FUNCTIONS ─────────────────────────────────────────────────

    def _xl(v):
        """Convert pandas NA/NaT to None so openpyxl can write the cell."""
        try:
            if pd.isna(v):
                return None
        except (TypeError, ValueError):
            pass
        return v

    def _col_fmt(col_name):
        """Return the Excel number format for a metric column."""
        if "Variance" in col_name or "Pct" in col_name or col_name.startswith("%") or "MAPE" in col_name:
            return _PCT_FMT
        if col_name in ("Volume", "Units Sold"):
            return "#,##0"
        if col_name in ("breakout", "period"):
            return "General"
        return _DOLLAR_FMT

    def _update_summary(ws, df):
        """Overwrite Summary sheet: write headers to row 1, clear rows 2+, paste all data."""
        ws.delete_rows(1, ws.max_row)

        df_cols = list(df.columns)
        bold = Font(bold=True)
        col_fills = [_col_fill(c) for c in df_cols]

        for ci, col_name in enumerate(df_cols, start=1):
            cell = ws.cell(1, ci)
            cell.value = col_name
            cell.font  = bold
            if col_fills[ci - 1]:
                cell.fill = col_fills[ci - 1]

        for i, (_, row_data) in enumerate(df.iterrows()):
            excel_row = 2 + i
            for ci, col_name in enumerate(df_cols, start=1):
                cell = ws.cell(excel_row, ci)
                cell.value         = _xl(row_data[col_name])
                cell.number_format = _col_fmt(col_name)

        print(f"  Summary done – {len(df):,} rows written.")

    def _write_sub_table(ws, df, start_row):
        """Write one sub-table (ACV Bucket / Loss Type / etc.) with period subheaders.

        Returns the first row after all written content.
        """
        df_cols     = list(df.columns)   # [period, bucket_col, metric1, …]
        period_col  = df_cols[0]
        bucket_col  = df_cols[1]
        metric_cols = df_cols[2:]
        met_fills   = [_col_fill(mc) for mc in metric_cols]

        bold = Font(bold=True)
        row  = start_row

        # Table header row
        ws.cell(row, 2).value = bucket_col
        ws.cell(row, 2).font  = bold
        for ci, mc in enumerate(metric_cols, start=3):
            cell = ws.cell(row, ci)
            cell.value = mc
            cell.font  = bold
            if met_fills[ci - 3]:
                cell.fill = met_fills[ci - 3]
        row += 1

        # One group per period
        for period in _PERIODS:
            period_df = df[df[period_col] == period]
            if period_df.empty:
                continue

            ws.cell(row, 1).value = period
            ws.cell(row, 1).font  = bold
            row += 1

            for _, row_data in period_df.iterrows():
                ws.cell(row, 2).value = _xl(row_data[bucket_col])
                for ci, mc in enumerate(metric_cols, start=3):
                    cell               = ws.cell(row, ci)
                    cell.value         = _xl(row_data[mc])
                    cell.number_format = _col_fmt(mc)
                row += 1

        return row

    def _update_detail_sheet(ws, tables):
        """Clear rows 3+ and write all sub-tables with period subheaders."""
        if ws.max_row >= 3:
            ws.delete_rows(3, ws.max_row - 2)

        current_row = 3
        for idx, df in enumerate(tables):
            current_row = _write_sub_table(ws, df, current_row)
            if idx < len(tables) - 1:
                current_row += 1  # blank gap row between sub-tables

        print(f"  Detail sheet done – last data row: {current_row - 1}.")

    # ── BUILD OUTPUT WORKBOOK ──────────────────────────────────────────────────

    output_filename = f"{pwd}/output/{job_name}/{job_name}_{current_run_date}.{output_file_type}"
    shutil.copy2(TEMPLATE_PATH, output_filename)
    print(f"Template copied → {output_filename}")

    wb = load_workbook(output_filename)

    for old_name, new_name in _SHEET_RENAMES.items():
        if old_name in wb.sheetnames:
            wb[old_name].title = new_name

    print(f"Updating '{SUMMARY_SHEET}'...")
    _update_summary(wb[SUMMARY_SHEET], df2)

    print(f"Updating '{BLUCAR_SHEET}'...")
    _update_detail_sheet(wb[BLUCAR_SHEET], [
        dfs["PQ.ai Error ACV Bucket View - BluCar ex-TFSS"],
        dfs["PQ.ai Error Sale Price Bucket View - BluCar ex-TFSS"],
        dfs["PQ.ai Error Title Type - BluCar ex-TFSS"],
        dfs["PQ.ai Error AutoGrade Bucket - BluCar ex-TFSS"],
        dfs["PQ.ai Error Loss Type - BluCar ex-TFSS"],
        dfs["PQ.ai Error Lot Type - BluCar ex-TFSS"],
    ])

    print(f"Updating '{INS_TFSS_SHEET}'...")
    _update_detail_sheet(wb[INS_TFSS_SHEET], [
        dfs["PQ.ai Error ACV Bucket View - Insurance + TFSS"],
        dfs["PQ.ai Error Sale Price Bucket View - Insurance + TFSS"],
        dfs["PQ.ai Error Title Type - Insurance + TFSS"],
        dfs["PQ.ai Error Make Bucket - Insurance + TFSS"],
        dfs["PQ.ai Error Loss Type - Insurance + TFSS"],
        dfs["PQ.ai Error Lot Type - Insurance + TFSS"],
    ])

    for ws in wb.worksheets:
        # Reset to a single fresh Selection — strips stale bottomLeft/bottomRight
        # entries inherited from the template (which would trigger Excel's "View"
        # repair against the new xSplit-only pane), while leaving the one entry
        # that openpyxl's freeze_panes setter expects to populate.
        ws.sheet_view.selection = [Selection()]
        ws.freeze_panes = "C1"
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            col_letter = get_column_letter(col_idx)
            for row_idx in range(1, min(ws.max_row + 1, 50)):
                val = ws.cell(row_idx, col_idx).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

    wb.save(output_filename)
    wb.close()
    print(f"Workbook saved.")

    # openpyxl rewrites xl/drawings/drawing1.xml with a default namespace
    # instead of the template's xdr:* prefix, which triggers Excel's
    # "Repaired Records: Drawing shape" on open. Restore the part
    # byte-for-byte from the template, preserving per-entry compression.
    _DRAWING_PARTS = ("xl/drawings/drawing1.xml",)
    with zipfile.ZipFile(TEMPLATE_PATH, "r") as tz:
        template_blobs = {p: tz.read(p) for p in _DRAWING_PARTS if p in tz.namelist()}
    if template_blobs:
        patched_path = output_filename + ".patched"
        with zipfile.ZipFile(output_filename, "r") as zin, \
             zipfile.ZipFile(patched_path, "w") as zout:
            for item in zin.infolist():
                data = template_blobs.get(item.filename, zin.read(item.filename))
                zout.writestr(item, data)
        shutil.move(patched_path, output_filename)
        print("Drawing part restored from template.")

    remove_old_files(f"{pwd}/output/{job_name}", file_type=output_file_type, keep=keep)
    print(f"job {job_name} completed on {datetime.datetime.now()}")

    return output_filename


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--job_name", default='')
    parser.add_argument("--current_run_date", default='')
    parser.add_argument("--key_path", required=True, help="Path to GCP service-account JSON key")
    args = parser.parse_args()

    result_path = run_report(
        job_name=args.job_name,
        current_run_date=args.current_run_date,
        key_path=args.key_path,
    )
    print(f"Excel report saved to: {result_path}")